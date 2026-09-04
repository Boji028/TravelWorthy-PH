"""Admin panel routes for site management and content control."""
from typing import Optional
from datetime import datetime, timezone
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, jsonify
from flask_login import current_user
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
import bleach
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy import or_, func
import io
from app import db
from decorators import admin_required
from utils import delete_old_image, save_image_metadata
from models.user import User
from models.package import TourPackage
from models.package_image import PackageImage
from models.inquiry import Inquiry
from models.blog import BlogPost
from constants import InquiryStatus
from image_service import ImageUploadService, ImageUploadException
from models.continent import Continent
from models.country import Country
from models.visa import VisaCountry
from models.testimonial import Testimonial
from models.inquiry_notification import InquiryNotification
from models.site_settings import SiteSettings
from models.email_verification import EmailVerificationToken
from models.agent import Agent
from models.itinerary_day import ItineraryDay
from models.travel_date import TravelDate
from models.hero_slide import HeroSlide

admin_bp = Blueprint("admin", __name__)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

# Fix 16: Explicit tag + attribute allowlist — blocks javascript: hrefs
ALLOWED_BLOG_TAGS = ["b", "i", "u", "em", "strong", "p", "br", "ul", "ol", "li", "h2", "h3", "a", "blockquote"]
ALLOWED_BLOG_ATTRS = {"a": ["href", "title"]}


def allowed_file(filename: str) -> bool:
    """Check if file extension is allowed.

    Args:
        filename: The filename to check

    Returns:
        True if allowed, False otherwise
    """
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_image(file, prefix: str = "img") -> Optional[str]:
    """Validate, save and compress an uploaded image. Returns filename or None.

    Deprecated: Use ImageUploadService.upload_and_compress() instead.

    Args:
        file: FileStorage object
        prefix: Prefix for filename

    Returns:
        Image path or None on error
    """
    try:
        result = ImageUploadService.upload_and_compress(file, prefix)
        return result["path"] if isinstance(result, dict) else result
    except ImageUploadException:
        return None


def get_dashboard_stats():
    """Get dashboard statistics with optimized queries.

    Returns:
        Dictionary with dashboard metrics
    """
    try:
        total_users = User.query.filter_by(is_admin=False).count()
        total_packages = TourPackage.query.count()
        total_inquiries = Inquiry.query.count()
        in_progress_inquiries = Inquiry.query.filter_by(status=InquiryStatus.CONTACTED.value).count()
        new_inquiries = Inquiry.query.filter_by(status=InquiryStatus.NEW.value).count()

        # FIX: joinedload prevents N+1 queries when the template accesses inquiry.package.title
        recent_inquiries = (
            Inquiry.query.options(joinedload(Inquiry.package)).order_by(Inquiry.created_at.desc()).limit(5).all()
        )

        # 7-day inquiry trend for the dashboard chart
        from datetime import timedelta

        today = datetime.now(timezone.utc).date()
        week_start = today - timedelta(days=6)
        week_start_dt = datetime.combine(week_start, datetime.min.time(), tzinfo=timezone.utc)

        week_inquiries = Inquiry.query.filter(Inquiry.created_at >= week_start_dt).all()

        daily_counts = {week_start + timedelta(days=i): 0 for i in range(7)}
        for inq in week_inquiries:
            idate = inq.created_at.date()
            if idate in daily_counts:
                daily_counts[idate] += 1

        inquiry_trend = [
            {"label": (week_start + timedelta(days=i)).strftime("%a"), "count": daily_counts[week_start + timedelta(days=i)]}
            for i in range(7)
        ]
        max_trend_count = max((d["count"] for d in inquiry_trend), default=0)

        return {
            "total_users": total_users,
            "total_packages": total_packages,
            "total_inquiries": total_inquiries,
            "in_progress_inquiries": in_progress_inquiries,
            "new_inquiries": new_inquiries,
            "recent_inquiries": recent_inquiries,
            "inquiry_trend": inquiry_trend,
            "max_trend_count": max_trend_count,
        }
    except Exception as e:
        current_app.logger.error(f"Dashboard stats error: {e}", exc_info=True)
        return {
            "total_users": 0,
            "total_packages": 0,
            "total_inquiries": 0,
            "in_progress_inquiries": 0,
            "new_inquiries": 0,
            "recent_inquiries": [],
            "inquiry_trend": [],
            "max_trend_count": 0,
        }


# ── Dashboard ──────────────────────────────────────────────
@admin_bp.route("/")
@admin_required
def dashboard() -> str:
    """Admin dashboard with site statistics."""
    stats = get_dashboard_stats()
    return render_template("admin/dashboard.html", **stats)


@admin_bp.route("/countries-by-continent/<int:continent_id>")
@admin_required
def countries_by_continent(continent_id):
    countries = Country.query.filter_by(continent_id=continent_id, is_active=True).order_by(Country.name).all()
    return jsonify([{"id": c.id, "name": c.name, "flag_emoji": c.flag_emoji} for c in countries])


# ── Packages ───────────────────────────────────────────────
@admin_bp.route("/packages")
@admin_required
def packages():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "").strip()
    continent_id = request.args.get("continent_id", type=int)
    status_filter = request.args.get("status", "").strip()

    query = TourPackage.query

    if search:
        query = query.filter(or_(TourPackage.title.ilike(f"%{search}%"), TourPackage.destination.ilike(f"%{search}%")))

    if continent_id:
        query = query.filter(TourPackage.country_id.in_(db.session.query(Country.id).filter_by(continent_id=continent_id)))

    if status_filter == "active":
        query = query.filter_by(is_active=True)
    elif status_filter == "inactive":
        query = query.filter_by(is_active=False)
    elif status_filter == "featured":
        query = query.filter_by(is_featured=True)

    all_packages = query.order_by(TourPackage.created_at.desc()).paginate(page=page, per_page=20, error_out=False)

    # Efficient inquiry-count per package — one aggregate query, not N+1
    package_ids = [p.id for p in all_packages.items]
    package_inquiry_counts = {}
    if package_ids:
        stats_rows = (
            db.session.query(Inquiry.package_id, func.count(Inquiry.id).label("inquiry_count"))
            .filter(Inquiry.package_id.in_(package_ids))
            .group_by(Inquiry.package_id)
            .all()
        )
        package_inquiry_counts = {row.package_id: row.inquiry_count for row in stats_rows}

    continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()

    return render_template(
        "admin/packages.html",
        packages=all_packages.items,
        pagination=all_packages,
        package_inquiry_counts=package_inquiry_counts,
        continents=continents,
        search=search,
        continent_id=continent_id,
        status_filter=status_filter,
    )


@admin_bp.route("/packages/toggle-active/<int:package_id>", methods=["POST"])
@admin_required
def toggle_package_active(package_id):
    package = db.get_or_404(TourPackage, package_id)
    package.is_active = not package.is_active
    db.session.commit()
    return jsonify(success=True, is_active=package.is_active)


@admin_bp.route("/packages/toggle-featured/<int:package_id>", methods=["POST"])
@admin_required
def toggle_package_featured(package_id):
    package = db.get_or_404(TourPackage, package_id)
    package.is_featured = not package.is_featured
    db.session.commit()
    return jsonify(success=True, is_featured=package.is_featured)


@admin_bp.route("/packages/bulk-action", methods=["POST"])
@admin_required
def bulk_package_action():
    action = request.form.get("action")
    package_ids = request.form.getlist("package_ids", type=int)

    if not package_ids:
        flash("No packages selected.", "warning")
        return redirect(url_for("admin.packages"))

    if action == "activate":
        TourPackage.query.filter(TourPackage.id.in_(package_ids)).update({"is_active": True}, synchronize_session=False)
        db.session.commit()
        flash(f"{len(package_ids)} package(s) activated.", "success")

    elif action == "deactivate":
        TourPackage.query.filter(TourPackage.id.in_(package_ids)).update({"is_active": False}, synchronize_session=False)
        db.session.commit()
        flash(f"{len(package_ids)} package(s) deactivated.", "success")

    elif action == "delete":
        packages_to_check = TourPackage.query.filter(TourPackage.id.in_(package_ids)).all()
        deleted, skipped = 0, 0
        for pkg in packages_to_check:
            open_inquiries = Inquiry.query.filter(
                Inquiry.package_id == pkg.id, Inquiry.status != InquiryStatus.CLOSED.value
            ).count()
            if open_inquiries > 0:
                skipped += 1
                continue
            delete_old_image(pkg.image, current_app.config["UPLOAD_FOLDER"])
            delete_old_image(pkg.flier_image, current_app.config["UPLOAD_FOLDER"])
            for img in pkg.images:
                delete_old_image(img.path, current_app.config["UPLOAD_FOLDER"])
            db.session.delete(pkg)
            deleted += 1
        db.session.commit()
        if skipped:
            flash(f"{deleted} package(s) deleted. {skipped} skipped (open inquiries exist).", "warning")
        else:
            flash(f"{deleted} package(s) deleted.", "success")
    else:
        flash("Invalid bulk action.", "danger")

    return redirect(url_for("admin.packages"))


@admin_bp.route("/packages/add", methods=["GET", "POST"])
@admin_required
def add_package():
    """Add new tour package with comprehensive error handling."""
    if request.method == "POST":
        try:
            title = request.form.get("title", "").strip()
            description = request.form.get("description", "").strip()
            destination = request.form.get("destination", "").strip()
            currency = request.form.get("currency", "PHP")

            if not all([title, description, destination]):
                flash("Please fill in all required fields.", "danger")
                continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                return render_template(
                    "admin/add_package.html",
                    continents=continents,
                    agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                )

            # Wrap all numeric casts in try/except
            try:
                country_id = int(request.form.get("country_id")) if request.form.get("country_id") else None
                duration_days = int(request.form.get("duration_days", 1))
                price = float(request.form.get("price", 0))
                latitude = float(request.form.get("latitude")) if request.form.get("latitude") else None
                longitude = float(request.form.get("longitude")) if request.form.get("longitude") else None

                if duration_days <= 0 or price < 0:
                    flash("Duration and price must be positive values.", "danger")
                    continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                    return render_template(
                        "admin/add_package.html",
                        continents=continents,
                        agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                    )
            except (ValueError, TypeError):
                flash("Invalid numeric values. Please check duration and price.", "danger")
                continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                return render_template(
                    "admin/add_package.html",
                    continents=continents,
                    agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                )

            filename = "default_tour.jpg"
            upload_result = None
            image_url = request.form.get("image_url", "").strip()
            if image_url and image_url.startswith("https://"):
                filename = image_url
            else:
                try:
                    image_file = request.files.get("image")
                    if image_file and image_file.filename:
                        upload_result = ImageUploadService.upload_and_compress(image_file, "package")
                        filename = upload_result["path"]
                except ImageUploadException as e:
                    current_app.logger.warning(f"Package image upload failed: {e}")
                    flash(f"Image upload failed: {str(e)}. Using default image.", "warning")
                except Exception as e:
                    current_app.logger.error(f"Unexpected error uploading package image: {e}", exc_info=True)
                    flash("Image upload error. Using default image.", "warning")

            # ── Flier image (optional) ──────────────────────────────
            flier_filename = None
            flier_upload_result = None
            flier_image_url = request.form.get("flier_image_url", "").strip()
            if flier_image_url and flier_image_url.startswith("https://"):
                flier_filename = flier_image_url
            else:
                try:
                    flier_file = request.files.get("flier_image")
                    if flier_file and flier_file.filename:
                        flier_upload_result = ImageUploadService.upload_and_compress(flier_file, "flier")
                        flier_filename = flier_upload_result["path"]
                except ImageUploadException as e:
                    current_app.logger.warning(f"Flier image upload failed: {e}")
                    flash(f"Flier upload failed: {str(e)}. You can add it later via Edit Package.", "warning")
                except Exception as e:
                    current_app.logger.error(f"Unexpected error uploading flier: {e}", exc_info=True)
                    flash("Flier upload error. You can add it later via Edit Package.", "warning")

            is_featured = request.form.get("is_featured") == "on"
            package_type = request.form.get("package_type", "domestic")
            if package_type not in ("domestic", "international"):
                package_type = "domestic"
            raw_agent_id = request.form.get("assigned_agent_id")
            assigned_agent_id = int(raw_agent_id) if raw_agent_id else None
            if assigned_agent_id and not db.session.get(Agent, assigned_agent_id):
                assigned_agent_id = None
            package = TourPackage(
                title=title,
                description=description,
                destination=destination,
                country_id=country_id,
                duration_days=duration_days,
                price=price,
                currency=currency,
                image=filename,
                flier_image=flier_filename,
                inclusions=request.form.get("inclusions", "").strip(),
                exclusions=request.form.get("exclusions", "").strip(),
                amenities=request.form.get("amenities", "").strip(),
                location_description=request.form.get("location_description", "").strip(),
                latitude=latitude,
                longitude=longitude,
                is_featured=is_featured,
                package_type=package_type,
                assigned_agent_id=assigned_agent_id,
            )
            db.session.add(package)
            db.session.commit()

            day_titles = request.form.getlist("itinerary_day_title")
            day_meals = request.form.getlist("itinerary_day_meals")
            day_descriptions = request.form.getlist("itinerary_day_description")
            day_images = _resolve_itinerary_day_images(len(day_titles))
            for i, day_title in enumerate(day_titles):
                day_title = day_title.strip()
                if not day_title:
                    continue
                db.session.add(
                    ItineraryDay(
                        package_id=package.id,
                        day_number=i + 1,
                        title=day_title,
                        meals=day_meals[i].strip() if i < len(day_meals) else None,
                        description=day_descriptions[i].strip() if i < len(day_descriptions) else None,
                        image=day_images[i] if i < len(day_images) else None,
                        order=i,
                    )
                )

            travel_date_values = request.form.getlist("travel_date_date")
            travel_date_end_values = request.form.getlist("travel_date_end_date")
            travel_date_notes = request.form.getlist("travel_date_note")
            for i, raw_date in enumerate(travel_date_values):
                raw_date = raw_date.strip()
                if not raw_date:
                    continue
                try:
                    parsed_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
                except ValueError:
                    continue
                parsed_end_date = None
                raw_end_date = travel_date_end_values[i].strip() if i < len(travel_date_end_values) else ""
                if raw_end_date:
                    try:
                        parsed_end_date = datetime.strptime(raw_end_date, "%Y-%m-%d").date()
                    except ValueError:
                        parsed_end_date = None
                db.session.add(
                    TravelDate(
                        package_id=package.id,
                        date=parsed_date,
                        end_date=parsed_end_date,
                        note=travel_date_notes[i].strip() if i < len(travel_date_notes) else None,
                    )
                )

            new_gallery_urls = request.form.getlist("new_gallery_urls")
            for order, url in enumerate(new_gallery_urls):
                if url and url.startswith("https://"):
                    db.session.add(PackageImage(package_id=package.id, path=url, order=order))
            db.session.commit()

            try:
                from notification_service import notify_users_new_package

                notify_users_new_package(package)
                db.session.commit()
            except Exception as notif_err:
                db.session.rollback()
                current_app.logger.warning(f"New-package notification failed for package #{package.id}: {notif_err}", exc_info=True)

            if upload_result:
                try:
                    save_image_metadata(package, upload_result, field_prefix="image")
                    db.session.commit()
                except Exception as e:
                    current_app.logger.warning(f"Could not save image metadata for package {package.id}: {e}")
            if flier_upload_result:
                try:
                    package.flier_image_size_kb = flier_upload_result.get("size_kb")
                    package.flier_image_uploaded_at = flier_upload_result.get("uploaded_at")
                    db.session.commit()
                except Exception as e:
                    current_app.logger.warning(f"Could not save flier metadata for package {package.id}: {e}")
            current_app.logger.info(f"Package added by admin: id={package.id}, title={title}")
            flash("Tour package added successfully!", "success")
            return redirect(url_for("admin.packages"))

        except IntegrityError as e:
            db.session.rollback()
            current_app.logger.error(f"Database integrity error adding package: {e}", exc_info=True)
            flash("A package with this title may already exist. Please try another.", "danger")
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error adding package: {e}", exc_info=True)
            flash("Database error occurred. Please try again.", "danger")
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Unexpected error adding package: {e}", exc_info=True)
            flash("An unexpected error occurred. Please try again.", "danger")

        continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
        return render_template(
            "admin/add_package.html",
            continents=continents,
            agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
        )

    continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
    return render_template(
        "admin/add_package.html",
        continents=continents,
        agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
    )


@admin_bp.route("/packages/edit/<int:package_id>", methods=["GET", "POST"])
@admin_required
def edit_package(package_id):
    """Edit tour package with comprehensive error handling."""
    try:
        package = db.get_or_404(TourPackage, package_id)
    except Exception as e:
        current_app.logger.error(f"Error loading package {package_id}: {e}", exc_info=True)
        flash("Package not found.", "danger")
        return redirect(url_for("admin.packages"))

    if request.method == "POST":
        try:
            # Capture into locals first — don't touch the model until all validation passes
            title = request.form.get("title", "").strip()
            description = request.form.get("description", "").strip()
            destination = request.form.get("destination", "").strip()

            if not all([title, description, destination]):
                flash("Please fill in all required fields.", "danger")
                continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                return render_template(
                    "admin/edit_package.html",
                    package=package,
                    continents=continents,
                    agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                )

            # Validate numerics before touching the model
            try:
                duration_days = int(request.form.get("duration_days", 1))
                price = float(request.form.get("price", 0))
                country_id = int(request.form.get("country_id")) if request.form.get("country_id") else None

                if duration_days <= 0 or price < 0:
                    flash("Duration and price must be positive values.", "danger")
                    continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                    return render_template(
                        "admin/edit_package.html",
                        package=package,
                        continents=continents,
                        agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                    )
            except (ValueError, TypeError):
                flash("Invalid numeric values. Please check duration and price.", "danger")
                continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
                return render_template(
                    "admin/edit_package.html",
                    package=package,
                    continents=continents,
                    agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
                )

            # All validation passed — now safe to assign to the model
            package.title = title
            package.description = description
            package.destination = destination
            package.duration_days = duration_days
            package.price = price
            package.country_id = country_id
            package.is_active = request.form.get("is_active") == "on"
            package.is_featured = request.form.get("is_featured") == "on"
            package.inclusions = request.form.get("inclusions", "").strip()
            package.exclusions = request.form.get("exclusions", "").strip()
            package.amenities = request.form.get("amenities", "").strip()

            ItineraryDay.query.filter_by(package_id=package.id).delete()
            day_titles = request.form.getlist("itinerary_day_title")
            day_meals = request.form.getlist("itinerary_day_meals")
            day_descriptions = request.form.getlist("itinerary_day_description")
            day_images = _resolve_itinerary_day_images(len(day_titles))
            for i, day_title in enumerate(day_titles):
                day_title = day_title.strip()
                if not day_title:
                    continue
                db.session.add(
                    ItineraryDay(
                        package_id=package.id,
                        day_number=i + 1,
                        title=day_title,
                        meals=day_meals[i].strip() if i < len(day_meals) else None,
                        description=day_descriptions[i].strip() if i < len(day_descriptions) else None,
                        image=day_images[i] if i < len(day_images) else None,
                        order=i,
                    )
                )

            TravelDate.query.filter_by(package_id=package.id).delete()
            travel_date_values = request.form.getlist("travel_date_date")
            travel_date_end_values = request.form.getlist("travel_date_end_date")
            travel_date_notes = request.form.getlist("travel_date_note")
            for i, raw_date in enumerate(travel_date_values):
                raw_date = raw_date.strip()
                if not raw_date:
                    continue
                try:
                    parsed_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
                except ValueError:
                    continue
                parsed_end_date = None
                raw_end_date = travel_date_end_values[i].strip() if i < len(travel_date_end_values) else ""
                if raw_end_date:
                    try:
                        parsed_end_date = datetime.strptime(raw_end_date, "%Y-%m-%d").date()
                    except ValueError:
                        parsed_end_date = None
                db.session.add(
                    TravelDate(
                        package_id=package.id,
                        date=parsed_date,
                        end_date=parsed_end_date,
                        note=travel_date_notes[i].strip() if i < len(travel_date_notes) else None,
                    )
                )

            package.location_description = request.form.get("location_description", "").strip()
            package.latitude = float(request.form.get("latitude")) if request.form.get("latitude") else None
            package.longitude = float(request.form.get("longitude")) if request.form.get("longitude") else None
            package.currency = request.form.get("currency", "PHP")
            package.package_type = (
                request.form.get("package_type")
                if request.form.get("package_type") in ("domestic", "international")
                else "domestic"
            )
            raw_agent_id = request.form.get("assigned_agent_id")
            agent_id_val = int(raw_agent_id) if raw_agent_id else None
            package.assigned_agent_id = agent_id_val if (agent_id_val and db.session.get(Agent, agent_id_val)) else None

            try:
                image_file = request.files.get("image")
                if image_file and image_file.filename:
                    upload_result = ImageUploadService.upload_and_compress(image_file, "package")
                    delete_old_image(package.image, current_app.config["UPLOAD_FOLDER"])
                    package.image = upload_result["path"]
                    save_image_metadata(package, upload_result, field_prefix="image")
            except ImageUploadException as e:
                current_app.logger.warning(f"Package image update failed: {e}")
                flash(f"Image upload failed: {str(e)}", "warning")
            except Exception as e:
                current_app.logger.error(f"Unexpected error updating package image: {e}", exc_info=True)
                flash("Image update error occurred.", "warning")
            # ── Flier image ──────────────────────────────────────────
            if request.form.get("remove_flier_image"):
                if package.flier_image:
                    try:
                        ImageUploadService.delete_image(package.flier_image)
                    except Exception as e:
                        current_app.logger.warning(f"Could not delete flier from storage: {e}")
                package.flier_image = None
                package.flier_image_size_kb = None
                package.flier_image_uploaded_at = None
            else:
                try:
                    flier_file = request.files.get("flier_image")
                    if flier_file and flier_file.filename:
                        if package.flier_image:
                            try:
                                ImageUploadService.delete_image(package.flier_image)
                            except Exception as e:
                                current_app.logger.warning(f"Could not delete old flier: {e}")
                        flier_result = ImageUploadService.upload_and_compress(flier_file, "flier")
                        package.flier_image = flier_result["path"]
                        package.flier_image_size_kb = flier_result.get("size_kb")
                        package.flier_image_uploaded_at = flier_result.get("uploaded_at")
                except ImageUploadException as e:
                    current_app.logger.warning(f"Flier update failed: {e}")
                    flash(f"Flier upload failed: {str(e)}", "warning")
                except Exception as e:
                    current_app.logger.error(f"Unexpected error updating flier: {e}", exc_info=True)
                    flash("Flier update error occurred.", "warning")
            # Track order manually: package.images does not include rows added
            # in this request until flush, so len() alone would give every new
            # image the same order value.
            next_order = len(package.images)

            # Handle direct Cloudinary upload URLs
            new_gallery_urls = request.form.getlist("new_gallery_urls")
            for url in new_gallery_urls:
                if url and url.startswith("https://"):
                    gallery_img = PackageImage(package_id=package.id, path=url, order=next_order)
                    db.session.add(gallery_img)
                    next_order += 1

            # Handle traditional file uploads (fallback)
            gallery_files = request.files.getlist("gallery_images")
            for gfile in gallery_files:
                if gfile and gfile.filename:
                    try:
                        upload_result = ImageUploadService.upload_and_compress(gfile, "package")
                        gallery_img = PackageImage(
                            package_id=package.id, path=upload_result["path"], order=next_order
                        )
                        db.session.add(gallery_img)
                        next_order += 1
                    except ImageUploadException as e:
                        flash(f"One image failed to upload: {str(e)}", "warning")

            db.session.commit()
            current_app.logger.info(f"Package updated by admin: id={package_id}")
            flash("Package updated successfully!", "success")
            return redirect(url_for("admin.packages"))

        except IntegrityError as e:
            db.session.rollback()
            current_app.logger.error(f"Database integrity error updating package {package_id}: {e}", exc_info=True)
            flash("Database integrity error. Please try again.", "danger")
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error updating package {package_id}: {e}", exc_info=True)
            flash("Database error occurred. Please try again.", "danger")
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Unexpected error updating package {package_id}: {e}", exc_info=True)
            flash("An unexpected error occurred. Please try again.", "danger")

    continents = Continent.query.filter_by(is_active=True).order_by(Continent.name).all()
    return render_template(
        "admin/edit_package.html",
        package=package,
        continents=continents,
        agents=Agent.query.filter_by(is_active=True).order_by(Agent.name).all(),
    )


@admin_bp.route("/packages/delete/<int:package_id>", methods=["POST"])
@admin_required
def delete_package(package_id):
    package = db.get_or_404(TourPackage, package_id)
    open_inquiries = Inquiry.query.filter(
        Inquiry.package_id == package_id, Inquiry.status != InquiryStatus.CLOSED.value
    ).count()
    if open_inquiries > 0:
        flash(
            f'Cannot delete — {open_inquiries} open inquir{"y" if open_inquiries == 1 else "ies"} exist for this package. Close them first.',
            "danger",
        )
        return redirect(url_for("admin.packages"))
    delete_old_image(package.image, current_app.config["UPLOAD_FOLDER"])
    delete_old_image(package.flier_image, current_app.config["UPLOAD_FOLDER"])
    for img in package.images:
        delete_old_image(img.path, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(package)
    db.session.commit()
    flash("Package deleted.", "info")
    return redirect(url_for("admin.packages"))


# ── Users ──────────────────────────────────────────────────
@admin_bp.route("/users")
@admin_required
def users():
    search = request.args.get("search", "").strip()
    role_filter = request.args.get("role", "").strip()

    query = User.query
    if search:
        query = query.filter(or_(User.name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%")))
    if role_filter == "admin":
        query = query.filter_by(is_admin=True)
    elif role_filter == "customer":
        query = query.filter_by(is_admin=False)

    page = request.args.get("page", 1, type=int)
    all_users = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20, error_out=False)

    # Match by email rather than user_id: guest inquiries have user_id=None,
    # so email is the only link that counts a user's pre-registration inquiries too.
    emails = [u.email for u in all_users.items]
    inquiry_counts_by_email = {}
    if emails:
        rows = (
            db.session.query(Inquiry.email, func.count(Inquiry.id))
            .filter(Inquiry.email.in_(emails))
            .group_by(Inquiry.email)
            .all()
        )
        inquiry_counts_by_email = dict(rows)

    return render_template(
        "admin/users.html",
        users=all_users.items,
        pagination=all_users,
        inquiry_counts_by_email=inquiry_counts_by_email,
        search=search,
        role_filter=role_filter,
    )


@admin_bp.route("/users/toggle-admin/<int:user_id>", methods=["POST"])
@admin_required
def toggle_user_admin(user_id):
    user = db.get_or_404(User, user_id)

    if user.id == current_user.id:
        flash("You cannot change your own admin status.", "danger")
        return redirect(url_for("admin.users"))

    user.is_admin = not user.is_admin
    db.session.flush()
    # Post-change check is atomic: if two concurrent requests both try to
    # remove the last admin, one will see 0 remaining after flush and rollback.
    if not user.is_admin and User.query.filter_by(is_admin=True).count() == 0:
        db.session.rollback()
        flash("Cannot remove admin — at least one admin account must remain.", "danger")
        return redirect(url_for("admin.users"))
    db.session.commit()
    flash(f'{user.name} is now {"an admin" if user.is_admin else "a customer"}.', "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    user = db.get_or_404(User, user_id)

    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin.users"))

    if user.is_admin:
        other_admins = User.query.filter(User.is_admin == True, User.id != user.id).count()
        if other_admins == 0:
            flash("Cannot delete — at least one admin account must remain.", "danger")
            return redirect(url_for("admin.users"))

    testimonial_count = Testimonial.query.filter_by(user_id=user.id).count()
    if testimonial_count > 0:
        flash(
            f"Cannot delete {user.name} — they have {testimonial_count} "
            f'testimonial{"s" if testimonial_count != 1 else ""}. Remove it from the Reviews admin first.',
            "danger",
        )
        return redirect(url_for("admin.users"))

    # Verification tokens are purely functional with no content worth preserving.
    EmailVerificationToken.query.filter_by(user_id=user.id).delete()
    # PackageReview.user now has cascade='all, delete-orphan' configured, so
    # any reviews by this user are removed automatically when the user is deleted.
    InquiryNotification.query.filter_by(user_id=user.id).delete()

    db.session.delete(user)
    db.session.commit()
    flash(f"{user.name} deleted.", "info")
    return redirect(url_for("admin.users"))


# ── Inquiries ──────────────────────────────────────────────
def _resolve_itinerary_day_images(day_count: int) -> list:
    """Work out the image URL for each submitted itinerary day.

    Both the add and edit forms post one file input plus one hidden
    "existing" input per day, so the lists line up with the day rows by
    index. A day keeps its current photo unless a new file is chosen —
    this matters because edit_package() deletes and recreates every
    ItineraryDay row on save, so without carrying the existing URL
    forward, editing a package's text would silently wipe its day photos.

    Returns a list of URLs (or None) indexed to match the day rows.
    """
    files = request.files.getlist("itinerary_day_image")
    existing = request.form.getlist("itinerary_day_image_existing")
    out = []
    for i in range(day_count):
        # An untouched file input still posts, with an empty filename.
        upload = files[i] if i < len(files) else None
        if upload and upload.filename:
            try:
                out.append(ImageUploadService.upload_and_compress(upload, "itinerary")["path"])
                continue
            except ImageUploadException as exc:
                flash(f"Day {i + 1} photo could not be uploaded: {exc}", "warning")
        prior = existing[i].strip() if i < len(existing) else ""
        out.append(prior or None)
    return out


def _inquiry_type(inq) -> str:
    """Classify an inquiry as 'package', 'visa', or 'trip'.

    There's no dedicated type column — package inquiries carry a
    package_id, visa inquiries are tagged with a '[FOR VISA]' prefix
    on special_requests, and everything else came from the Plan My
    Trip form.
    """
    if inq.package_id:
        return "package"
    if inq.special_requests and inq.special_requests.startswith("[FOR VISA]"):
        return "visa"
    return "trip"


def _get_inquiry_filter_params() -> dict:
    """Read and normalize the inquiry filter query params.

    Shared by the inquiries list view and the export route so they always
    agree on what "the current filters" means.
    """
    sort = request.args.get("sort", "desc").strip()
    # Default to current month if no date filter is set, to limit default query size.
    month_default = (
        datetime.now(timezone.utc).strftime("%Y-%m")
        if not any(request.args.get(k) for k in ("month", "year", "date_from", "date_to", "search", "status", "type"))
        else ""
    )
    return {
        "status": request.args.get("status", "").strip(),
        "type": request.args.get("type", "").strip(),
        "search": request.args.get("search", "").strip(),
        "month": request.args.get("month", month_default).strip(),
        "year": request.args.get("year", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "sort": sort if sort in ("asc", "desc") else "desc",
    }


def _apply_inquiry_filters(query, params: dict):
    """Apply type/search/date-range filters. Status is applied separately."""
    from datetime import timedelta

    if params["type"] == "package":
        query = query.filter(Inquiry.package_id.isnot(None))
    elif params["type"] == "visa":
        query = query.filter(Inquiry.package_id.is_(None), Inquiry.special_requests.like("[FOR VISA]%"))
    elif params["type"] == "trip":
        query = query.filter(
            Inquiry.package_id.is_(None),
            or_(Inquiry.special_requests.is_(None), ~Inquiry.special_requests.like("[FOR VISA]%")),
        )

    if params["search"]:
        query = query.filter(
            or_(
                Inquiry.name.ilike(f"%{params['search']}%"),
                Inquiry.email.ilike(f"%{params['search']}%"),
                Inquiry.reference_number.ilike(f"%{params['search']}%"),
            )
        )

    # Precedence: month quick-pick > year quick-pick > manual from/to range
    effective_from = None
    effective_to_exclusive = None
    if params["month"]:
        try:
            yr, mo = (int(p) for p in params["month"].split("-"))
            effective_from = datetime(yr, mo, 1, tzinfo=timezone.utc)
            effective_to_exclusive = (
                datetime(yr + 1, 1, 1, tzinfo=timezone.utc) if mo == 12 else datetime(yr, mo + 1, 1, tzinfo=timezone.utc)
            )
        except (ValueError, AttributeError):
            pass
    elif params["year"]:
        try:
            yr = int(params["year"])
            effective_from = datetime(yr, 1, 1, tzinfo=timezone.utc)
            effective_to_exclusive = datetime(yr + 1, 1, 1, tzinfo=timezone.utc)
        except ValueError:
            pass
    else:
        if params["date_from"]:
            try:
                effective_from = datetime.strptime(params["date_from"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                pass
        if params["date_to"]:
            try:
                effective_to_exclusive = datetime.strptime(params["date_to"], "%Y-%m-%d").replace(
                    tzinfo=timezone.utc
                ) + timedelta(days=1)
            except ValueError:
                pass

    if effective_from:
        query = query.filter(Inquiry.created_at >= effective_from)
    if effective_to_exclusive:
        query = query.filter(Inquiry.created_at < effective_to_exclusive)

    return query


@admin_bp.route("/inquiries")
@admin_required
def inquiries():
    params = _get_inquiry_filter_params()
    status_filter = params["status"]
    type_filter = params["type"]
    search = params["search"]
    month_param = params["month"]
    year_param = params["year"]
    date_from = params["date_from"]
    date_to = params["date_to"]
    sort = params["sort"]

    base_query = _apply_inquiry_filters(Inquiry.query, params)

    # Pill counts reflect the search/type/date filters but not the status pill itself,
    # so switching status pills doesn't make the other counts disappear.
    # Single query counts all statuses at once instead of one COUNT per status.
    from sqlalchemy import case

    count_rows = base_query.with_entities(
        func.count().label("total"),
        func.sum(case((Inquiry.status == "new", 1), else_=0)).label("new"),
        func.sum(case((Inquiry.status == "contacted", 1), else_=0)).label("contacted"),
        func.sum(case((Inquiry.status == "confirmed", 1), else_=0)).label("confirmed"),
        func.sum(case((Inquiry.status == "closed", 1), else_=0)).label("closed"),
    ).one()
    status_counts = {
        "all": count_rows.total or 0,
        "new": count_rows.new or 0,
        "contacted": count_rows.contacted or 0,
        "confirmed": count_rows.confirmed or 0,
        "closed": count_rows.closed or 0,
    }

    # joinedload: the template renders inq.package.title per row, which
    # otherwise lazy-loads one package query per inquiry on the page.
    query = base_query.options(joinedload(Inquiry.package))
    if status_filter:
        query = query.filter_by(status=status_filter)
    query = query.order_by(Inquiry.created_at.asc() if sort == "asc" else Inquiry.created_at.desc())

    page = request.args.get("page", 1, type=int)
    all_inquiries = query.paginate(page=page, per_page=20, error_out=False)

    # Quick-pick labels/values for the "All time / This month / Last month / This year" dropdown
    today = datetime.now(timezone.utc)
    this_month_value = today.strftime("%Y-%m")
    this_month_label = today.strftime("%b %Y")
    last_month_dt = (
        today.replace(year=today.year - 1, month=12, day=1)
        if today.month == 1
        else today.replace(month=today.month - 1, day=1)
    )
    last_month_value = last_month_dt.strftime("%Y-%m")
    last_month_label = last_month_dt.strftime("%b %Y")
    this_year_value = today.strftime("%Y")

    return render_template(
        "admin/inquiries.html",
        inquiries=all_inquiries.items,
        pagination=all_inquiries,
        status_filter=status_filter,
        type_filter=type_filter,
        search=search,
        month_param=month_param,
        year_param=year_param,
        date_from=date_from,
        date_to=date_to,
        sort=sort,
        status_counts=status_counts,
        this_month_value=this_month_value,
        this_month_label=this_month_label,
        last_month_value=last_month_value,
        last_month_label=last_month_label,
        this_year_value=this_year_value,
    )


@admin_bp.route("/inquiries/report")
@admin_required
def inquiry_report():
    """Printable summary of inquiries for a chosen month or year."""
    import calendar

    period_type = request.args.get("period_type", "month").strip()
    period_value = request.args.get("period_value", "").strip()
    today = datetime.now(timezone.utc)

    if period_type == "year":
        try:
            year = int(period_value) if period_value else today.year
        except ValueError:
            year = today.year
        start = datetime(year, 1, 1, tzinfo=timezone.utc)
        end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        period_label = str(year)
        period_value = str(year)
    else:
        period_type = "month"
        try:
            yr, mo = (int(p) for p in period_value.split("-"))
        except (ValueError, AttributeError):
            yr, mo = today.year, today.month
        start = datetime(yr, mo, 1, tzinfo=timezone.utc)
        end = datetime(yr + 1, 1, 1, tzinfo=timezone.utc) if mo == 12 else datetime(yr, mo + 1, 1, tzinfo=timezone.utc)
        period_label = f"{calendar.month_name[mo]} {yr}"
        period_value = f"{yr:04d}-{mo:02d}"

    report_inquiries = (
        Inquiry.query.filter(Inquiry.created_at >= start, Inquiry.created_at < end).order_by(Inquiry.created_at.desc()).all()
    )

    type_counts = {"package": 0, "visa": 0, "trip": 0}
    status_counts = {}
    for inq in report_inquiries:
        type_counts[_inquiry_type(inq)] += 1
        status_counts[inq.status] = status_counts.get(inq.status, 0) + 1

    return render_template(
        "admin/inquiry_report.html",
        inquiries=report_inquiries,
        period_label=period_label,
        period_type=period_type,
        period_value=period_value,
        type_counts=type_counts,
        status_counts=status_counts,
        total=len(report_inquiries),
        classify=_inquiry_type,
        generated_at=today,
    )


@admin_bp.route("/inquiries/update/<int:inquiry_id>", methods=["POST"])
@admin_required
def update_inquiry_status(inquiry_id):
    inquiry = db.get_or_404(Inquiry, inquiry_id)
    new_status = request.form.get("status")
    valid_inquiry_statuses = [s.value for s in InquiryStatus]
    try:
        if new_status not in valid_inquiry_statuses:
            flash(f'Invalid status. Choose from: {", ".join(valid_inquiry_statuses)}.', "danger")
            return redirect(url_for("admin.inquiries"))

        old_status = inquiry.status
        inquiry.status = new_status

        if inquiry.status == InquiryStatus.CONFIRMED.value and old_status != InquiryStatus.CONFIRMED.value:
            try:
                from email_service import send_inquiry_confirmed

                send_inquiry_confirmed(inquiry)
            except Exception as email_err:
                current_app.logger.warning(f"Confirmation email failed for inquiry #{inquiry.id}: {email_err}", exc_info=True)

        if new_status != old_status:
            from notification_service import notify_inquiry_status_change

            notify_inquiry_status_change(inquiry, f"Your inquiry to {inquiry.destination} is now {new_status}.")

        db.session.commit()
        flash(f"Inquiry #{inquiry.id} status updated to {new_status}.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating inquiry {inquiry_id}: {e}", exc_info=True)
        flash("Failed to update inquiry status.", "danger")
    return redirect(url_for("admin.inquiries"))


@admin_bp.route("/inquiries/reply/<int:inquiry_id>", methods=["POST"])
@admin_required
def reply_to_inquiry(inquiry_id):
    inquiry = db.get_or_404(Inquiry, inquiry_id)
    admin_response = request.form.get("response", "").strip()

    if not admin_response:
        flash("Please provide a response message.", "danger")
        return redirect(url_for("admin.inquiries"))

    try:
        from email_service import send_inquiry_reply

        # send_inquiry_reply's success/failure return value is not checked here
        # (unlike send_inquiry_receipt) — see docs/fix-silent-inquiry-email-failures.md.
        send_inquiry_reply(inquiry, admin_response)

        # DB update proceeds regardless of whether the email above actually sent.
        inquiry.admin_response = admin_response
        inquiry.responded_at = datetime.now(timezone.utc)
        # Only advance to CONTACTED; never downgrade an already-confirmed/closed inquiry.
        if inquiry.status in (InquiryStatus.NEW.value, InquiryStatus.CONTACTED.value):
            inquiry.status = InquiryStatus.CONTACTED.value

        from notification_service import notify_inquiry_status_change

        notify_inquiry_status_change(inquiry, f"We've replied to your inquiry about {inquiry.destination}.")

        db.session.commit()

        flash(f"Response sent to {inquiry.email}!", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Reply to inquiry {inquiry_id} failed: {e}", exc_info=True)
        flash("Failed to send reply. Check logs for details.", "danger")
    return redirect(url_for("admin.inquiries"))


@admin_bp.route("/inquiries/delete/<int:inquiry_id>", methods=["POST"])
@admin_required
def delete_inquiry(inquiry_id):
    inquiry = db.get_or_404(Inquiry, inquiry_id)
    db.session.delete(inquiry)
    db.session.commit()
    flash("Inquiry deleted.", "info")
    return redirect(url_for("admin.inquiries"))


# ── Blog ──────────────────────────────────────────────────
@admin_bp.route("/blog")
@admin_required
def blog():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "").strip()
    category_filter = request.args.get("category", "").strip()
    status_filter = request.args.get("status", "").strip()

    base_query = BlogPost.query
    if search:
        base_query = base_query.filter(or_(BlogPost.title.ilike(f"%{search}%"), BlogPost.author.ilike(f"%{search}%")))
    if category_filter:
        base_query = base_query.filter_by(category=category_filter)

    status_counts = {
        "all": base_query.count(),
        "published": base_query.filter_by(is_published=True).count(),
        "draft": base_query.filter_by(is_published=False).count(),
    }

    query = base_query
    if status_filter == "published":
        query = query.filter_by(is_published=True)
    elif status_filter == "draft":
        query = query.filter_by(is_published=False)

    all_categories = [
        c[0]
        for c in db.session.query(BlogPost.category)
        .filter(BlogPost.category.isnot(None))
        .distinct()
        .order_by(BlogPost.category)
        .all()
    ]

    posts = query.order_by(BlogPost.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template(
        "admin/blog.html",
        posts=posts.items,
        pagination=posts,
        search=search,
        category_filter=category_filter,
        status_filter=status_filter,
        status_counts=status_counts,
        all_categories=all_categories,
    )


@admin_bp.route("/blog/toggle-published/<int:post_id>", methods=["POST"])
@admin_required
def toggle_blog_published(post_id):
    post = db.get_or_404(BlogPost, post_id)
    post.is_published = not post.is_published
    db.session.commit()
    return jsonify(success=True, is_published=post.is_published)


@admin_bp.route("/blog/add", methods=["GET", "POST"])
@admin_required
def add_blog():
    if request.method == "POST":
        try:
            title = request.form.get("title", "").strip()
            author = request.form.get("author", "Admin").strip()
            category = request.form.get("category", "").strip()
            short_description = request.form.get("short_description", "").strip()
            # Fix 16: Pass explicit attributes to block javascript: hrefs
            content = bleach.clean(
                request.form.get("content", "").strip(), tags=ALLOWED_BLOG_TAGS, attributes=ALLOWED_BLOG_ATTRS, strip=True
            )
            is_published = request.form.get("is_published") == "on"

            if not title or not content:
                flash("Title and content are required.", "danger")
                return redirect(url_for("admin.add_blog"))

            featured_image = None
            upload_result = None
            try:
                featured_file = request.files.get("featured_image")
                if featured_file and featured_file.filename:
                    current_app.logger.info(f"Uploading file: {featured_file.filename}")
                    upload_result = ImageUploadService.upload_and_compress(featured_file, "blog")
                    featured_image = upload_result["path"]
                    current_app.logger.info(f"Upload successful: {featured_image}")
            except ImageUploadException as e:
                current_app.logger.error(f"ImageUploadException: {str(e)}")
                flash(f"Image upload failed: {str(e)}", "danger")
                return redirect(url_for("admin.add_blog"))
            except Exception as e:
                current_app.logger.error(f"Image upload error: {type(e).__name__}: {str(e)}", exc_info=True)
                flash(f"Image upload failed: {str(e)}", "danger")
                return redirect(url_for("admin.add_blog"))

            post = BlogPost(
                title=title,
                author=author,
                category=category,
                short_description=short_description,
                content=content,
                featured_image=featured_image,
                is_published=is_published,
            )
            db.session.add(post)
            db.session.commit()
            current_app.logger.info(f"Blog post created: {post.id}")

            if upload_result:
                try:
                    save_image_metadata(post, upload_result, field_prefix="featured_image")
                    db.session.commit()
                    current_app.logger.info(f"Image metadata saved for post {post.id}")
                except Exception as e:
                    current_app.logger.error(f"Error saving image metadata: {type(e).__name__}: {str(e)}", exc_info=True)
                    # Don't fail the whole operation just because metadata save failed

            flash("Blog post published!", "success")
            return redirect(url_for("admin.blog"))

        except Exception as e:
            current_app.logger.error(f"Unexpected error in add_blog: {type(e).__name__}: {str(e)}", exc_info=True)
            flash("An unexpected error occurred. Please check the logs for details.", "danger")
            return redirect(url_for("admin.add_blog"))

    return render_template("admin/add_blog.html")


@admin_bp.route("/blog/edit/<int:post_id>", methods=["GET", "POST"])
@admin_required
def edit_blog(post_id):
    post = db.get_or_404(BlogPost, post_id)
    if request.method == "POST":
        # Validate into locals first — don't touch the model until all checks pass
        title = request.form.get("title", "").strip()
        author = request.form.get("author", "Admin").strip()
        category = request.form.get("category", "").strip()
        short_description = request.form.get("short_description", "").strip()
        content = bleach.clean(
            request.form.get("content", "").strip(), tags=ALLOWED_BLOG_TAGS, attributes=ALLOWED_BLOG_ATTRS, strip=True
        )
        is_published = request.form.get("is_published") == "on"

        if not title or not content:
            flash("Title and content are required.", "danger")
            return redirect(url_for("admin.edit_blog", post_id=post_id))

        try:
            featured_file = request.files.get("featured_image")
            if featured_file and featured_file.filename:
                upload_result = ImageUploadService.upload_and_compress(featured_file, "blog")
                delete_old_image(post.featured_image, current_app.config["UPLOAD_FOLDER"])
                post.featured_image = upload_result["path"]
                save_image_metadata(post, upload_result, field_prefix="featured_image")
        except ImageUploadException as e:
            flash(f"Image upload failed: {str(e)}", "danger")
            return redirect(url_for("admin.edit_blog", post_id=post_id))

        # All checks passed — now safe to assign to the model
        post.title = title
        post.author = author
        post.category = category
        post.short_description = short_description
        post.content = content
        post.is_published = is_published

        db.session.commit()
        flash("Blog post updated!", "success")
        return redirect(url_for("admin.blog"))
    return render_template("admin/edit_blog.html", post=post)


@admin_bp.route("/blog/delete/<int:post_id>", methods=["POST"])
@admin_required
def delete_blog(post_id):
    post = db.get_or_404(BlogPost, post_id)
    delete_old_image(post.featured_image, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(post)
    db.session.commit()
    flash("Blog post deleted.", "info")
    return redirect(url_for("admin.blog"))


# ── Continents ────────────────────────────────────────────
@admin_bp.route("/continents")
@admin_required
def continents():
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()

    query = Continent.query
    if search:
        query = query.filter(Continent.name.ilike(f"%{search}%"))
    if status_filter == "active":
        query = query.filter_by(is_active=True)
    elif status_filter == "inactive":
        query = query.filter_by(is_active=False)

    all_continents = query.order_by(Continent.name).all()

    # Efficient country-count + package-count per continent — two aggregate queries, not N+1
    country_counts = dict(db.session.query(Country.continent_id, func.count(Country.id)).group_by(Country.continent_id).all())
    package_counts = dict(
        db.session.query(Country.continent_id, func.count(TourPackage.id))
        .join(TourPackage, TourPackage.country_id == Country.id)
        .group_by(Country.continent_id)
        .all()
    )
    continent_stats = {
        c.id: {"country_count": country_counts.get(c.id, 0), "package_count": package_counts.get(c.id, 0)}
        for c in all_continents
    }

    return render_template(
        "admin/continents.html",
        continents=all_continents,
        continent_stats=continent_stats,
        search=search,
        status_filter=status_filter,
    )


@admin_bp.route("/continents/toggle-active/<int:continent_id>", methods=["POST"])
@admin_required
def toggle_continent_active(continent_id):
    continent = db.get_or_404(Continent, continent_id)
    continent.is_active = not continent.is_active
    db.session.commit()
    return jsonify(success=True, is_active=continent.is_active)


@admin_bp.route("/continents/add", methods=["GET", "POST"])
@admin_required
def add_continent():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        flag_emoji = request.form.get("flag_emoji", "").strip()
        description = request.form.get("description", "").strip()
        is_active = request.form.get("is_active") == "on"
        if not name:
            flash("Continent name is required.", "danger")
            return redirect(url_for("admin.add_continent"))

        image = ""
        upload_result = None
        try:
            image_file = request.files.get("image")
            if image_file and image_file.filename:
                upload_result = ImageUploadService.upload_and_compress(image_file, "continent")
                image = upload_result["path"]
        except ImageUploadException as e:
            flash(f"Image upload failed: {str(e)}", "danger")
            return redirect(url_for("admin.add_continent"))

        continent = Continent(name=name, flag_emoji=flag_emoji, description=description, image=image, is_active=is_active)
        db.session.add(continent)
        db.session.commit()
        if upload_result:
            save_image_metadata(continent, upload_result, field_prefix="image")
            db.session.commit()
        flash(f"{name} added successfully!", "success")
        return redirect(url_for("admin.continents"))
    return render_template("admin/add_continent.html")


@admin_bp.route("/continents/edit/<int:continent_id>", methods=["GET", "POST"])
@admin_required
def edit_continent(continent_id):
    continent = db.get_or_404(Continent, continent_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Continent name is required.", "danger")
            return redirect(url_for("admin.edit_continent", continent_id=continent_id))
        continent.name = name
        continent.flag_emoji = request.form.get("flag_emoji", "").strip()
        continent.description = request.form.get("description", "").strip()
        continent.is_active = request.form.get("is_active") == "on"

        try:
            image_file = request.files.get("image")
            if image_file and image_file.filename:
                upload_result = ImageUploadService.upload_and_compress(image_file, "continent")
                delete_old_image(continent.image, current_app.config["UPLOAD_FOLDER"])
                continent.image = upload_result["path"]
                save_image_metadata(continent, upload_result, field_prefix="image")
        except ImageUploadException as e:
            flash(f"Image upload failed: {str(e)}", "danger")
            return redirect(url_for("admin.edit_continent", continent_id=continent_id))

        db.session.commit()
        flash(f"{continent.name} updated!", "success")
        return redirect(url_for("admin.continents"))
    return render_template("admin/edit_continent.html", continent=continent)


@admin_bp.route("/continents/remove-image/<int:continent_id>", methods=["POST"])
@admin_required
def remove_continent_image(continent_id):
    """Remove the region image from a continent."""
    continent = db.get_or_404(Continent, continent_id)

    if not continent.image:
        flash("This region has no image to remove.", "warning")
        return redirect(url_for("admin.edit_continent", continent_id=continent_id))

    delete_old_image(continent.image, current_app.config["UPLOAD_FOLDER"])
    continent.image = None
    continent.image_size_kb = None
    continent.image_uploaded_at = None
    db.session.commit()
    flash("Region image removed.", "info")
    return redirect(url_for("admin.edit_continent", continent_id=continent_id))


@admin_bp.route("/continents/delete/<int:continent_id>", methods=["POST"])
@admin_required
def delete_continent(continent_id):
    continent = db.get_or_404(Continent, continent_id)
    country_count = Country.query.filter_by(continent_id=continent_id).count()
    if country_count > 0:
        flash(
            f"Cannot delete — {country_count} country/countries are assigned to this region. "
            "Reassign or delete them first.",
            "danger",
        )
        return redirect(url_for("admin.continents"))
    db.session.delete(continent)
    db.session.commit()
    flash("Continent deleted.", "info")
    return redirect(url_for("admin.continents"))


# ── Countries ─────────────────────────────────────────────
@admin_bp.route("/countries")
@admin_required
def countries():
    continent_id = request.args.get("continent_id", type=int)
    continent = db.get_or_404(Continent, continent_id) if continent_id else None
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()

    all_countries = []
    package_counts = {}
    if continent_id:
        query = Country.query.filter_by(continent_id=continent_id)
        if search:
            query = query.filter(Country.name.ilike(f"%{search}%"))
        if status_filter == "active":
            query = query.filter_by(is_active=True)
        elif status_filter == "inactive":
            query = query.filter_by(is_active=False)
        all_countries = query.order_by(Country.name).all()

        country_ids = [c.id for c in all_countries]
        if country_ids:
            package_counts = dict(
                db.session.query(TourPackage.country_id, func.count(TourPackage.id))
                .filter(TourPackage.country_id.in_(country_ids))
                .group_by(TourPackage.country_id)
                .all()
            )

    return render_template(
        "admin/countries.html",
        countries=all_countries,
        continent=continent,
        package_counts=package_counts,
        search=search,
        status_filter=status_filter,
    )


@admin_bp.route("/countries/toggle-active/<int:country_id>", methods=["POST"])
@admin_required
def toggle_country_active(country_id):
    country = db.get_or_404(Country, country_id)
    country.is_active = not country.is_active
    db.session.commit()
    return jsonify(success=True, is_active=country.is_active)


@admin_bp.route("/countries/add", methods=["GET", "POST"])
@admin_required
def add_country():
    continent_id = request.args.get("continent_id", type=int)
    continent = db.get_or_404(Continent, continent_id) if continent_id else None
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        flag_emoji = request.form.get("flag_emoji", "").strip()
        description = request.form.get("description", "").strip()
        is_active = request.form.get("is_active") == "on"
        if not name:
            flash("Country name is required.", "danger")
            return redirect(url_for("admin.countries", continent_id=continent_id))

        image = ""
        upload_result = None
        try:
            image_file = request.files.get("image")
            if image_file and image_file.filename:
                upload_result = ImageUploadService.upload_and_compress(image_file, "country")
                image = upload_result["path"]
        except ImageUploadException as e:
            flash(f"Image upload failed: {str(e)}", "danger")
            return redirect(url_for("admin.add_country", continent_id=continent_id))

        country = Country(
            name=name,
            flag_emoji=flag_emoji,
            description=description,
            image=image,
            is_active=is_active,
            continent_id=continent_id,
        )
        db.session.add(country)
        db.session.commit()
        if upload_result:
            save_image_metadata(country, upload_result, field_prefix="image")
            db.session.commit()
        flash(f"{name} added successfully!", "success")
        return redirect(url_for("admin.countries", continent_id=continent_id))
    return render_template("admin/add_country.html", continent=continent)


@admin_bp.route("/countries/edit/<int:country_id>", methods=["GET", "POST"])
@admin_required
def edit_country(country_id):
    country = db.get_or_404(Country, country_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Country name is required.", "danger")
            return redirect(url_for("admin.edit_country", country_id=country_id))
        country.name = name
        country.flag_emoji = request.form.get("flag_emoji", "").strip()
        country.description = request.form.get("description", "").strip()
        country.is_active = request.form.get("is_active") == "on"

        db.session.commit()

        try:
            image_file = request.files.get("image")
            if image_file and image_file.filename:
                upload_result = ImageUploadService.upload_and_compress(image_file, "country")
                delete_old_image(country.image, current_app.config["UPLOAD_FOLDER"])
                country.image = upload_result["path"]
                save_image_metadata(country, upload_result, field_prefix="image")
                db.session.commit()
        except ImageUploadException as e:
            flash(f"Country updated, but image upload failed: {str(e)}", "warning")
            return redirect(url_for("admin.countries", continent_id=country.continent_id))

        flash(f"{country.name} updated!", "success")
        return redirect(url_for("admin.countries", continent_id=country.continent_id))
    return render_template("admin/edit_country.html", country=country)


@admin_bp.route("/countries/remove-image/<int:country_id>", methods=["POST"])
@admin_required
def remove_country_image(country_id):
    """Remove the photo from a country."""
    country = db.get_or_404(Country, country_id)

    if not country.image:
        flash("This country has no image to remove.", "warning")
        return redirect(url_for("admin.edit_country", country_id=country_id))

    delete_old_image(country.image, current_app.config["UPLOAD_FOLDER"])
    country.image = None
    country.image_size_kb = None
    country.image_uploaded_at = None
    db.session.commit()
    flash("Country image removed.", "info")
    return redirect(url_for("admin.edit_country", country_id=country_id))


@admin_bp.route("/countries/delete/<int:country_id>", methods=["POST"])
@admin_required
def delete_country(country_id):
    country = db.get_or_404(Country, country_id)
    continent_id = country.continent_id
    package_count = TourPackage.query.filter_by(country_id=country.id).count()
    if package_count > 0:
        flash(
            f"Cannot delete {country.name} — it still has {package_count} "
            f'{"package" if package_count == 1 else "packages"} assigned. '
            f"Reassign or delete those packages first.",
            "danger",
        )
        return redirect(url_for("admin.countries", continent_id=continent_id))

    delete_old_image(country.image, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(country)
    db.session.commit()
    flash("Country deleted.", "info")
    return redirect(url_for("admin.countries", continent_id=continent_id))


# ── Visa ──────────────────────────────────────────────────
@admin_bp.route("/visa")
@admin_required
def visa_list():
    search = request.args.get("search", "").strip()
    region_filter = request.args.get("region", "").strip()
    status_filter = request.args.get("status", "").strip()

    base_query = VisaCountry.query
    if search:
        base_query = base_query.filter(VisaCountry.country_name.ilike(f"%{search}%"))
    if region_filter:
        base_query = base_query.filter_by(region=region_filter)

    status_counts = {
        "all": base_query.count(),
        "active": base_query.filter_by(is_active=True).count(),
        "inactive": base_query.filter_by(is_active=False).count(),
    }

    query = base_query
    if status_filter == "active":
        query = query.filter_by(is_active=True)
    elif status_filter == "inactive":
        query = query.filter_by(is_active=False)

    all_regions = [
        r[0]
        for r in db.session.query(VisaCountry.region)
        .filter(VisaCountry.region.isnot(None))
        .distinct()
        .order_by(VisaCountry.region)
        .all()
    ]

    visas = query.order_by(VisaCountry.country_name).all()
    return render_template(
        "admin/visa.html",
        visas=visas,
        search=search,
        region_filter=region_filter,
        status_filter=status_filter,
        status_counts=status_counts,
        all_regions=all_regions,
    )


@admin_bp.route("/visa/toggle-active/<int:visa_id>", methods=["POST"])
@admin_required
def toggle_visa_active(visa_id):
    visa = db.get_or_404(VisaCountry, visa_id)
    visa.is_active = not visa.is_active
    db.session.commit()
    return jsonify(success=True, is_active=visa.is_active)


@admin_bp.route("/visa/add", methods=["GET", "POST"])
@admin_required
def visa_add():
    if request.method == "POST":
        country_name = request.form.get("country_name", "").strip()
        flag_emoji = request.form.get("flag_emoji", "").strip()
        is_active = request.form.get("is_active") == "on"
        if not country_name:
            flash("Country name is required.", "danger")
            return redirect(url_for("admin.visa_add"))

        pdf_url = None
        pdf_file = request.files.get("requirements_pdf")
        if pdf_file and pdf_file.filename:
            try:
                upload_result = ImageUploadService.upload_pdf(pdf_file, "visa")
                pdf_url = upload_result["path"]
            except ImageUploadException as e:
                flash(f"PDF upload failed: {str(e)}", "danger")
                return redirect(url_for("admin.visa_add"))

        try:
            price = float(request.form.get("price")) if request.form.get("price") else None
        except ValueError:
            price = None
        if price is not None and price < 0:
            flash("Price cannot be negative.", "danger")
            return redirect(url_for("admin.visa_add"))

        region = request.form.get("region", "").strip() or None
        visa_type = request.form.get("visa_type", "").strip() or None
        processing_time = request.form.get("processing_time", "").strip() or None
        stay_validity = request.form.get("stay_validity", "").strip() or None
        try:
            documents_count = int(request.form.get("documents_count")) if request.form.get("documents_count") else None
        except ValueError:
            documents_count = None
        if documents_count is not None and documents_count < 0:
            flash("Documents count cannot be negative.", "danger")
            return redirect(url_for("admin.visa_add"))

        visa = VisaCountry(
            country_name=country_name,
            flag_emoji=flag_emoji,
            requirements_pdf=pdf_url,
            price=price,
            is_active=is_active,
            region=region,
            visa_type=visa_type,
            processing_time=processing_time,
            stay_validity=stay_validity,
            documents_count=documents_count,
        )
        db.session.add(visa)
        db.session.commit()

        try:
            from notification_service import notify_users_new_visa

            notify_users_new_visa(visa)
            db.session.commit()
        except Exception as notif_err:
            db.session.rollback()
            current_app.logger.warning(f"New-visa notification failed for visa #{visa.id}: {notif_err}", exc_info=True)

        flash(f"{country_name} visa added!", "success")
        return redirect(url_for("admin.visa_list"))
    return render_template("admin/add_visa.html")


@admin_bp.route("/visa/edit/<int:visa_id>", methods=["GET", "POST"])
@admin_required
def visa_edit(visa_id):
    visa = db.get_or_404(VisaCountry, visa_id)
    if request.method == "POST":
        country_name = request.form.get("country_name", "").strip()
        if not country_name:
            flash("Country name is required.", "danger")
            return redirect(url_for("admin.visa_edit", visa_id=visa_id))

        # Validate numerics before mutating any model fields or touching files.
        try:
            new_price = float(request.form.get("price")) if request.form.get("price") else None
        except ValueError:
            new_price = None
        if new_price is not None and new_price < 0:
            flash("Price cannot be negative.", "danger")
            return redirect(url_for("admin.visa_edit", visa_id=visa_id))

        try:
            new_docs = int(request.form.get("documents_count")) if request.form.get("documents_count") else None
        except ValueError:
            new_docs = None
        if new_docs is not None and new_docs < 0:
            flash("Documents count cannot be negative.", "danger")
            return redirect(url_for("admin.visa_edit", visa_id=visa_id))

        # PDF replacement runs last: it deletes the old file from storage, so any
        # validation failure after this point would leave the DB pointing at a
        # file that no longer exists.
        pdf_file = request.files.get("requirements_pdf")
        if pdf_file and pdf_file.filename:
            try:
                upload_result = ImageUploadService.upload_pdf(pdf_file, "visa")
                delete_old_image(visa.requirements_pdf, current_app.config["UPLOAD_FOLDER"])
                visa.requirements_pdf = upload_result["path"]
            except ImageUploadException as e:
                flash(f"PDF upload failed: {str(e)}", "danger")
                return redirect(url_for("admin.visa_edit", visa_id=visa_id))

        # All validation passed — now mutate the model.
        visa.country_name = country_name
        visa.flag_emoji = request.form.get("flag_emoji", "").strip()
        visa.is_active = request.form.get("is_active") == "on"
        visa.price = new_price
        visa.region = request.form.get("region", "").strip() or None
        visa.visa_type = request.form.get("visa_type", "").strip() or None
        visa.processing_time = request.form.get("processing_time", "").strip() or None
        visa.stay_validity = request.form.get("stay_validity", "").strip() or None
        visa.documents_count = new_docs

        db.session.commit()
        flash(f"{visa.country_name} updated!", "success")
        return redirect(url_for("admin.visa_list"))
    return render_template("admin/edit_visa.html", visa=visa)


@admin_bp.route("/visa/delete/<int:visa_id>", methods=["POST"])
@admin_required
def visa_delete(visa_id):
    visa = db.get_or_404(VisaCountry, visa_id)
    delete_old_image(visa.requirements_pdf, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(visa)
    db.session.commit()
    flash("Visa entry deleted.", "info")
    return redirect(url_for("admin.visa_list"))


# ── Testimonials ──────────────────────────────────────────
@admin_bp.route("/testimonials")
@admin_required
def testimonials():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "").strip()
    rating_filter = request.args.get("rating", "").strip()

    base_query = Testimonial.query.join(User, Testimonial.user_id == User.id)
    if search:
        base_query = base_query.filter(or_(User.name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%")))

    rating_counts = {"all": base_query.count()}
    for r in range(1, 6):
        rating_counts[str(r)] = base_query.filter(Testimonial.rating == r).count()

    query = base_query
    if rating_filter in ("1", "2", "3", "4", "5"):
        query = query.filter(Testimonial.rating == int(rating_filter))

    # selectinload: the template renders testimonial.user.name/.email and
    # testimonial.images per row (same pattern as main.reviews — see
    # docs/testimonial-review-selectinload-n-plus-one.md).
    testimonials_data = (
        query.options(selectinload(Testimonial.user), selectinload(Testimonial.images))
        .order_by(Testimonial.created_at.desc())
        .paginate(page=page, per_page=20, error_out=False)
    )
    return render_template(
        "admin/testimonials.html",
        testimonials=testimonials_data.items,
        pagination=testimonials_data,
        search=search,
        rating_filter=rating_filter,
        rating_counts=rating_counts,
    )


@admin_bp.route("/testimonials/delete-photo/<int:testimonial_id>", methods=["POST"])
@admin_required
def delete_testimonial_photo(testimonial_id):
    testimonial = db.get_or_404(Testimonial, testimonial_id)
    from models.testimonial_image import TestimonialImage

    images = TestimonialImage.query.filter_by(testimonial_id=testimonial.id).all()
    if images:
        for img in images:
            delete_old_image(img.path, current_app.config["UPLOAD_FOLDER"])
            db.session.delete(img)
        db.session.commit()
        flash("Testimonial photo(s) removed.", "success")
    return redirect(url_for("admin.testimonials"))


# ── Photo Removal Routes (for quick deletion) ──────────────
@admin_bp.route("/packages/remove-photo/<int:package_id>", methods=["POST"])
@admin_required
def remove_package_photo(package_id):
    """Remove photo from a package."""
    package = db.get_or_404(TourPackage, package_id)
    if package.image and package.image != "default_tour.jpg":
        delete_old_image(package.image, current_app.config["UPLOAD_FOLDER"])
        package.image = None
        db.session.commit()
        flash("Package photo removed.", "success")
    else:
        flash("No photo to remove.", "warning")
    return redirect(url_for("admin.edit_package", package_id=package_id))


@admin_bp.route("/blog/remove-photo/<int:post_id>", methods=["POST"])
@admin_required
def remove_blog_photo(post_id):
    """Remove featured image from a blog post."""
    post = db.get_or_404(BlogPost, post_id)
    if post.featured_image:
        delete_old_image(post.featured_image, current_app.config["UPLOAD_FOLDER"])
        post.featured_image = None
        db.session.commit()
        flash("Blog photo removed.", "success")
    else:
        flash("No photo to remove.", "warning")
    return redirect(url_for("admin.edit_blog", post_id=post_id))


@admin_bp.route("/visa/remove-pdf/<int:visa_id>", methods=["POST"])
@admin_required
def remove_visa_pdf(visa_id):
    """Remove requirements PDF from a visa entry."""
    visa = db.get_or_404(VisaCountry, visa_id)
    if visa.requirements_pdf:
        delete_old_image(visa.requirements_pdf, current_app.config["UPLOAD_FOLDER"])
        visa.requirements_pdf = None
        db.session.commit()
        flash("Visa PDF removed.", "success")
    else:
        flash("No PDF to remove.", "warning")
    return redirect(url_for("admin.visa_edit", visa_id=visa_id))


@admin_bp.route("/packages/delete-gallery-image/<int:image_id>", methods=["POST"])
@admin_required
def delete_gallery_image(image_id):
    img = db.get_or_404(PackageImage, image_id)
    package_id = img.package_id
    delete_old_image(img.path, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(img)
    db.session.commit()
    flash("Gallery image deleted.", "info")
    return redirect(url_for("admin.edit_package", package_id=package_id))


@admin_bp.route("/cloudinary-signature", methods=["POST"])
@admin_required
def cloudinary_signature():
    """Generate a signed Cloudinary upload signature for direct browser uploads."""
    import cloudinary.utils
    import time

    timestamp = int(time.time())
    folder = request.json.get("folder", "travelworthyph/package")

    params_to_sign = {
        "timestamp": timestamp,
        "folder": folder,
    }

    signature = cloudinary.utils.api_sign_request(params_to_sign, cloudinary.config().api_secret)

    return jsonify(
        {
            "signature": signature,
            "timestamp": timestamp,
            "cloud_name": cloudinary.config().cloud_name,
            "api_key": cloudinary.config().api_key,
            "folder": folder,
        }
    )


# ── Excel Exports ──────────────────────────────────────────
@admin_bp.route("/inquiries/export")
@admin_required
def export_inquiries():
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    params = _get_inquiry_filter_params()
    # joinedload: each row writes inq.package.title, and the export set is
    # unbounded — without this, exporting N inquiries fires N lazy-load queries.
    query = _apply_inquiry_filters(Inquiry.query.options(joinedload(Inquiry.package)), params)
    if params["status"]:
        query = query.filter_by(status=params["status"])
    query = query.order_by(Inquiry.created_at.asc() if params["sort"] == "asc" else Inquiry.created_at.desc())

    inquiries = query.all()

    if inquiries:
        Inquiry.query.filter(Inquiry.id.in_([inq.id for inq in inquiries])).update(
            {"last_exported_at": datetime.now(timezone.utc)}, synchronize_session=False
        )
        db.session.commit()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"

    headers = [
        "Reference",
        "Name",
        "Email",
        "Confirmation Email Failed",
        "Contact",
        "Destination",
        "Package",
        "Date From",
        "Date To",
        "Adults",
        "Children",
        "Infants",
        "Special Requests",
        "Status",
        "Admin Response",
        "Created At",
        "Responded At",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="175968", end_color="175968", fill_type="solid")
    header_font = Font(color="FDFAF6", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for inq in inquiries:
        ws.append(
            [
                inq.reference_number,
                inq.name,
                inq.email,
                "Yes" if inq.confirmation_email_failed else "",
                inq.contact_number,
                inq.destination,
                inq.package.title if inq.package else "",
                inq.travel_date_from,
                inq.travel_date_to,
                inq.num_adults,
                inq.num_children,
                inq.num_infants,
                inq.special_requests or "",
                inq.status,
                inq.admin_response or "",
                inq.created_at.replace(tzinfo=None) if inq.created_at else "",
                inq.responded_at.replace(tzinfo=None) if inq.responded_at else "",
            ]
        )

    # Excel date columns: dates as yyyy-mm-dd, timestamps with time included.
    # Column indices below account for "Confirmation Email Failed" being
    # inserted at position 4 — everything after it shifted by one.
    last_row = ws.max_row
    for col_idx in (8, 9):
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).number_format = "yyyy-mm-dd"
    for col_idx in (16, 17):
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).number_format = "yyyy-mm-dd hh:mm"

    widths = [12, 20, 26, 16, 14, 20, 22, 12, 12, 8, 9, 8, 30, 12, 30, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inquiries_{params['status'] or 'all'}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@admin_bp.route("/fliers/<int:package_id>/remove", methods=["POST"])
@admin_required
def remove_flier(package_id):
    """Remove the flier image from a package."""
    package = db.get_or_404(TourPackage, package_id)

    if not package.flier_image:
        flash("This package has no flier to remove.", "warning")
        return redirect(url_for("admin.packages"))

    try:
        ImageUploadService.delete_image(package.flier_image)
    except Exception as e:
        current_app.logger.warning(f"Could not delete flier from Cloudinary for package {package_id}: {e}")

    package.flier_image = None
    package.flier_image_size_kb = None
    package.flier_image_uploaded_at = None

    try:
        db.session.commit()
        flash(f'Flier for "{package.title}" removed.', "success")
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"DB error removing flier for package {package_id}: {e}", exc_info=True)
        flash("Database error removing flier. Please try again.", "danger")

    return redirect(url_for("admin.packages"))


def _site_image_payload(path, size_kb, uploaded_at):
    """Build the JSON-friendly representation of a site-settings image field."""
    if not path:
        return None
    return {
        "url": path if path.startswith("http") else f"/uploads/{path}",
        "size_kb": round(size_kb, 1) if size_kb else None,
        "uploaded_at": uploaded_at.strftime("%b %d, %Y") if uploaded_at else None,
    }


@admin_bp.route("/site-settings", methods=["GET", "POST"])
@admin_required
def site_settings():
    """Manage the admin-configurable homepage background images
    (hero, testimonials, closing CTA)."""
    settings = SiteSettings.get_settings()

    if request.method == "POST":
        fields = [
            ("hero_image", "site_hero"),
            ("testimonial_image", "site_testimonial"),
            ("cta_image", "site_cta"),
        ]
        for field_name, cloudinary_folder in fields:
            try:
                image_file = request.files.get(field_name)
                if image_file and image_file.filename:
                    upload_result = ImageUploadService.upload_and_compress(image_file, cloudinary_folder)
                    delete_old_image(getattr(settings, field_name), current_app.config["UPLOAD_FOLDER"])
                    setattr(settings, field_name, upload_result["path"])
                    save_image_metadata(settings, upload_result, field_prefix=field_name)
            except ImageUploadException as e:
                return jsonify(success=False, message=f'{field_name.replace("_", " ").title()} upload failed: {str(e)}'), 400

        db.session.commit()
        return jsonify(
            success=True,
            message="Site settings updated!",
            updated_at=settings.updated_at.strftime("%b %d, %Y") if settings.updated_at else None,
            images={
                field_name: _site_image_payload(
                    getattr(settings, field_name),
                    getattr(settings, f"{field_name}_size_kb"),
                    getattr(settings, f"{field_name}_uploaded_at"),
                )
                for field_name, _ in fields
            },
        )

    return render_template("admin/site_settings.html", settings=settings)

# ── Hero Slides ─────────────────────────────────────────
@admin_bp.route("/hero-slides")
@admin_required
def hero_slides():
    slides = HeroSlide.query.order_by(HeroSlide.order).all()
    return render_template("admin/hero_slides.html", slides=slides)


@admin_bp.route("/hero-slides/add", methods=["POST"])
@admin_required
def add_hero_slides():
    # Track order manually, same reasoning as the package gallery upload:
    # newly-added rows aren't visible via a fresh count() until committed,
    # so compute the starting point once, before adding anything.
    next_order = HeroSlide.query.count()
    new_urls = request.form.getlist("new_slide_urls")
    added = 0
    for url in new_urls:
        if url and url.startswith("https://"):
            db.session.add(HeroSlide(path=url, order=next_order))
            next_order += 1
            added += 1
    db.session.commit()
    if added:
        flash(f"{added} slide{'s' if added != 1 else ''} added.", "success")
    else:
        flash("No images were uploaded.", "warning")
    return redirect(url_for("admin.hero_slides"))


@admin_bp.route("/hero-slides/delete/<int:slide_id>", methods=["POST"])
@admin_required
def delete_hero_slide(slide_id):
    slide = db.get_or_404(HeroSlide, slide_id)
    delete_old_image(slide.path, current_app.config["UPLOAD_FOLDER"])
    db.session.delete(slide)
    db.session.commit()
    flash("Slide removed.", "info")
    return redirect(url_for("admin.hero_slides"))


@admin_bp.route("/hero-slides/reorder/<int:slide_id>/<direction>", methods=["POST"])
@admin_required
def reorder_hero_slide(slide_id, direction):
    slide = db.get_or_404(HeroSlide, slide_id)
    if direction == "up":
        neighbor = (
            HeroSlide.query.filter(HeroSlide.order < slide.order).order_by(HeroSlide.order.desc()).first()
        )
    elif direction == "down":
        neighbor = (
            HeroSlide.query.filter(HeroSlide.order > slide.order).order_by(HeroSlide.order.asc()).first()
        )
    else:
        neighbor = None
    if neighbor:
        slide.order, neighbor.order = neighbor.order, slide.order
        db.session.commit()
    return redirect(url_for("admin.hero_slides"))

@admin_bp.route("/hero-slides/set-mobile/<int:slide_id>", methods=["POST"])
@admin_required
def set_hero_slide_mobile(slide_id):
    """Attach a separately-composed portrait photo to one slide, for phone
    screens — an escape hatch from the automatic ar_9:16 smart-crop when a
    landscape crop of the same shot isn't the best framing."""
    slide = db.get_or_404(HeroSlide, slide_id)
    url = request.form.get("mobile_url", "").strip()
    if url and url.startswith("https://"):
        if slide.mobile_path:
            delete_old_image(slide.mobile_path, current_app.config["UPLOAD_FOLDER"])
        slide.mobile_path = url
        db.session.commit()
        flash("Mobile version added.", "success")
    else:
        flash("No image was uploaded.", "warning")
    return redirect(url_for("admin.hero_slides"))


@admin_bp.route("/hero-slides/remove-mobile/<int:slide_id>", methods=["POST"])
@admin_required
def remove_hero_slide_mobile(slide_id):
    slide = db.get_or_404(HeroSlide, slide_id)
    if slide.mobile_path:
        delete_old_image(slide.mobile_path, current_app.config["UPLOAD_FOLDER"])
        slide.mobile_path = None
        db.session.commit()
        flash("Mobile version removed — back to the automatic crop.", "info")
    return redirect(url_for("admin.hero_slides"))
# ── Agents ──────────────────────────────────────────────
@admin_bp.route("/agents")
@admin_required
def agents():
    all_agents = Agent.query.order_by(Agent.name).all()
    return render_template("admin/agents.html", agents=all_agents)


@admin_bp.route("/agents/add", methods=["GET", "POST"])
@admin_required
def add_agent():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        notes = request.form.get("notes", "").strip()
        is_active = request.form.get("is_active") == "on"
        if not name or not email:
            flash("Name and email are required.", "danger")
            return redirect(url_for("admin.add_agent"))
        agent = Agent(name=name, email=email, notes=notes or None, is_active=is_active)
        db.session.add(agent)
        db.session.commit()
        flash(f"{name} added as an agent.", "success")
        return redirect(url_for("admin.agents"))
    return render_template("admin/add_agent.html")


@admin_bp.route("/agents/edit/<int:agent_id>", methods=["GET", "POST"])
@admin_required
def edit_agent(agent_id):
    agent = db.get_or_404(Agent, agent_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        if not name or not email:
            flash("Agent name and email are required.", "danger")
            return redirect(url_for("admin.edit_agent", agent_id=agent_id))
        agent.name = name
        agent.email = email
        agent.notes = request.form.get("notes", "").strip() or None
        agent.is_active = request.form.get("is_active") == "on"
        db.session.commit()
        flash(f"{agent.name} updated.", "success")
        return redirect(url_for("admin.agents"))
    return render_template("admin/edit_agent.html", agent=agent)


@admin_bp.route("/agents/toggle-active/<int:agent_id>", methods=["POST"])
@admin_required
def toggle_agent_active(agent_id):
    agent = db.get_or_404(Agent, agent_id)
    agent.is_active = not agent.is_active
    db.session.commit()
    return jsonify(success=True, is_active=agent.is_active)


@admin_bp.route("/agents/set-visa-agent/<int:agent_id>", methods=["POST"])
@admin_required
def set_visa_agent(agent_id):
    """Mark one agent as the visa-inquiries agent — only one can hold this
    role at a time, so clear it from everyone else first."""
    agent = db.get_or_404(Agent, agent_id)
    Agent.query.filter(Agent.id != agent_id).update({"is_visa_agent": False}, synchronize_session=False)
    agent.is_visa_agent = True
    db.session.commit()
    return jsonify(success=True, agent_id=agent.id)


@admin_bp.route("/agents/unset-visa-agent/<int:agent_id>", methods=["POST"])
@admin_required
def unset_visa_agent(agent_id):
    agent = db.get_or_404(Agent, agent_id)
    agent.is_visa_agent = False
    db.session.commit()
    return jsonify(success=True)


@admin_bp.route("/agents/delete/<int:agent_id>", methods=["POST"])
@admin_required
def delete_agent(agent_id):
    agent = db.get_or_404(Agent, agent_id)
    assigned_count = TourPackage.query.filter_by(assigned_agent_id=agent.id).count()
    if assigned_count > 0:
        flash(
            f"Cannot delete {agent.name} — still assigned to {assigned_count} "
            f'package{"s" if assigned_count != 1 else ""}. Reassign those packages first.',
            "danger",
        )
        return redirect(url_for("admin.agents"))
    was_visa_agent = agent.is_visa_agent
    db.session.delete(agent)
    db.session.commit()
    if was_visa_agent:
        flash(
            f"{agent.name} deleted. They were the visa agent — visa inquiries "
            "will no longer be CC'd to anyone until you set a new one.",
            "warning",
        )
    else:
        flash(f"{agent.name} deleted.", "info")
    return redirect(url_for("admin.agents"))


@admin_bp.route("/site-settings/remove-image/<string:field>", methods=["POST"])
@admin_required
def remove_site_image(field):
    """Remove one of the admin-configurable homepage background images."""
    valid_fields = {"hero_image", "testimonial_image", "cta_image"}
    if field not in valid_fields:
        return jsonify(success=False, message="Invalid image field."), 400

    settings = SiteSettings.get_settings()
    current_image = getattr(settings, field)

    if not current_image:
        return jsonify(success=False, message="No image to remove."), 400

    delete_old_image(current_image, current_app.config["UPLOAD_FOLDER"])
    setattr(settings, field, None)
    setattr(settings, f"{field}_size_kb", None)
    setattr(settings, f"{field}_uploaded_at", None)
    db.session.commit()
    return jsonify(success=True, message="Image removed.", field=field)
