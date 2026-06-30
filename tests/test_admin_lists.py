"""Tests for admin list/filter pages and utility endpoints."""
from datetime import date, timedelta
from models.package import TourPackage
from models.inquiry import Inquiry
from models.continent import Continent
from models.country import Country
from models.user import User
from werkzeug.security import generate_password_hash


def _make_package(db, **overrides):
    defaults = dict(
        title='Tour Package',
        description='A tour.',
        destination='Palawan',
        duration_days=5,
        price=5000.00,
        currency='PHP',
        is_active=True,
        is_featured=False,
    )
    defaults.update(overrides)
    pkg = TourPackage(**defaults)
    db.session.add(pkg)
    db.session.commit()
    return pkg


def _make_inquiry(db, **overrides):
    today = date.today()
    defaults = dict(
        name='Customer',
        email='customer@example.com',
        contact_number='+639171234567',
        destination='Palawan',
        travel_date_from=today + timedelta(days=10),
        travel_date_to=today + timedelta(days=14),
        num_adults=2,
        status='new',
    )
    defaults.update(overrides)
    inquiry = Inquiry(**defaults)
    db.session.add(inquiry)
    db.session.commit()
    return inquiry


def _xlsx_strings(data: bytes) -> list:
    """Flatten all non-empty cell values from the export's first worksheet."""
    from openpyxl import load_workbook
    import io as _io
    wb = load_workbook(_io.BytesIO(data))
    ws = wb.active
    return [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]


def _make_user(db, **overrides):
    defaults = dict(
        name='User',
        email='user@example.com',
        password=generate_password_hash('Pass123'),
        is_admin=False,
        email_verified=True,
    )
    defaults.update(overrides)
    user = User(**defaults)
    db.session.add(user)
    db.session.commit()
    return user


class TestDashboard:
    def test_requires_login(self, client):
        response = client.get('/admin/')
        assert response.status_code in (302, 401, 403)

    def test_rejects_non_admin(self, app, authenticated_client):
        response = authenticated_client.get('/admin/')
        assert response.status_code in (302, 403)

    def test_admin_can_access(self, app, admin_client):
        response = admin_client.get('/admin/')
        assert response.status_code == 200

    def test_renders_with_data(self, app, admin_client):
        from app import db
        _make_package(db)
        _make_inquiry(db)
        response = admin_client.get('/admin/')
        assert response.status_code == 200


class TestPackagesList:
    def test_requires_login(self, client):
        response = client.get('/admin/packages')
        assert response.status_code in (302, 401, 403)

    def test_admin_can_access(self, app, admin_client):
        response = admin_client.get('/admin/packages')
        assert response.status_code == 200

    def test_search_filter(self, app, admin_client):
        from app import db
        _make_package(db, title='Palawan Adventure', destination='Palawan')
        _make_package(db, title='Cebu Island', destination='Cebu')
        response = admin_client.get('/admin/packages?search=Palawan')
        assert b'Palawan Adventure' in response.data
        assert b'Cebu Island' not in response.data

    def test_status_filter_active(self, app, admin_client):
        from app import db
        _make_package(db, title='ZZZActivePackage', is_active=True)
        _make_package(db, title='ZZZHiddenPackage', is_active=False)
        response = admin_client.get('/admin/packages?status=active')
        assert b'ZZZActivePackage' in response.data
        assert b'ZZZHiddenPackage' not in response.data

    def test_status_filter_featured(self, app, admin_client):
        from app import db
        _make_package(db, title='ZZZFeaturedPackage', is_featured=True)
        _make_package(db, title='ZZZRegularPackage', is_featured=False)
        response = admin_client.get('/admin/packages?status=featured')
        assert b'ZZZFeaturedPackage' in response.data
        assert b'ZZZRegularPackage' not in response.data


class TestUsersList:
    def test_requires_login(self, client):
        response = client.get('/admin/users')
        assert response.status_code in (302, 401, 403)

    def test_admin_can_access(self, app, admin_client):
        response = admin_client.get('/admin/users')
        assert response.status_code == 200

    def test_search_by_name(self, app, admin_client):
        from app import db
        _make_user(db, name='Juan dela Cruz', email='juan@example.com')
        _make_user(db, name='Maria Santos', email='maria@example.com')
        response = admin_client.get('/admin/users?search=Juan')
        assert b'Juan dela Cruz' in response.data
        assert b'Maria Santos' not in response.data

    def test_search_by_email(self, app, admin_client):
        from app import db
        _make_user(db, name='Test', email='findme@example.com')
        response = admin_client.get('/admin/users?search=findme')
        assert b'findme@example.com' in response.data

    def test_role_filter_admin(self, app, admin_client):
        from app import db
        _make_user(db, name='Regular', email='regular@example.com', is_admin=False)
        _make_user(db, name='OtherAdmin', email='otheradmin@example.com', is_admin=True)
        response = admin_client.get('/admin/users?role=admin')
        assert b'OtherAdmin' in response.data
        assert b'Regular' not in response.data

    def test_role_filter_customer(self, app, admin_client):
        from app import db
        _make_user(db, name='Customer', email='cust@example.com', is_admin=False)
        response = admin_client.get('/admin/users?role=customer')
        assert b'Customer' in response.data


class TestInquiriesList:
    def test_requires_login(self, client):
        response = client.get('/admin/inquiries')
        assert response.status_code in (302, 401, 403)

    def test_admin_can_access(self, app, admin_client):
        response = admin_client.get('/admin/inquiries')
        assert response.status_code == 200

    def test_renders_with_inquiries(self, app, admin_client):
        from app import db
        _make_inquiry(db)
        response = admin_client.get('/admin/inquiries')
        assert response.status_code == 200

    def test_status_filter(self, app, admin_client):
        from app import db
        _make_inquiry(db, email='new@example.com', status='new')
        _make_inquiry(db, email='closed@example.com', status='closed')
        response = admin_client.get('/admin/inquiries?status=new')
        assert response.status_code == 200


class TestCountriesByContinent:
    def test_requires_login(self, client):
        response = client.get('/admin/countries-by-continent/1')
        assert response.status_code in (302, 401, 403)

    def test_returns_json_list(self, app, admin_client):
        from app import db
        continent = Continent(name='Asia', is_active=True)
        db.session.add(continent)
        db.session.commit()
        country = Country(name='Japan', is_active=True, continent_id=continent.id)
        db.session.add(country)
        db.session.commit()
        response = admin_client.get(f'/admin/countries-by-continent/{continent.id}')
        assert response.status_code == 200
        data = response.get_json()
        assert any(c['name'] == 'Japan' for c in data)

    def test_returns_empty_list_for_empty_continent(self, app, admin_client):
        from app import db
        continent = Continent(name='Empty', is_active=True)
        db.session.add(continent)
        db.session.commit()
        response = admin_client.get(f'/admin/countries-by-continent/{continent.id}')
        assert response.get_json() == []

    def test_excludes_inactive_countries(self, app, admin_client):
        from app import db
        continent = Continent(name='Asia2', is_active=True)
        db.session.add(continent)
        db.session.commit()
        country = Country(name='Inactive Country', is_active=False, continent_id=continent.id)
        db.session.add(country)
        db.session.commit()
        response = admin_client.get(f'/admin/countries-by-continent/{continent.id}')
        assert response.get_json() == []


class TestExportInquiries:
    def test_requires_login(self, client):
        response = client.get('/admin/inquiries/export')
        assert response.status_code in (302, 401, 403)

    def test_returns_xlsx(self, app, admin_client):
        from app import db
        _make_inquiry(db)
        response = admin_client.get('/admin/inquiries/export')
        assert response.status_code == 200
        assert 'spreadsheetml.sheet' in response.content_type

    def test_header_row(self, app, admin_client):
        response = admin_client.get('/admin/inquiries/export')
        values = _xlsx_strings(response.data)
        assert 'Reference' in values
        assert 'Destination' in values

    def test_contains_inquiry_data(self, app, admin_client):
        from app import db
        _make_inquiry(db, destination='Boracay', email='test@export.com')
        response = admin_client.get('/admin/inquiries/export')
        assert 'Boracay' in _xlsx_strings(response.data)

    def test_status_filter_applied(self, app, admin_client):
        from app import db
        _make_inquiry(db, email='new@example.com', destination='New Dest', status='new')
        _make_inquiry(db, email='closed@example.com', destination='Closed Dest', status='closed')
        response = admin_client.get('/admin/inquiries/export?status=new')
        values = _xlsx_strings(response.data)
        assert 'New Dest' in values
        assert 'Closed Dest' not in values

    def test_search_filter_applied(self, app, admin_client):
        from app import db
        _make_inquiry(db, email='findme@example.com', destination='Searchable Dest')
        _make_inquiry(db, email='other@example.com', destination='Other Dest')
        response = admin_client.get('/admin/inquiries/export?search=findme')
        values = _xlsx_strings(response.data)
        assert 'Searchable Dest' in values
        assert 'Other Dest' not in values

    def test_type_filter_applied(self, app, admin_client):
        from app import db
        _make_inquiry(db, email='trip@example.com', destination='Trip Dest')
        _make_inquiry(db, email='visa@example.com', destination='Visa Dest', special_requests='[FOR VISA] Japan')
        response = admin_client.get('/admin/inquiries/export?type=visa')
        values = _xlsx_strings(response.data)
        assert 'Visa Dest' in values
        assert 'Trip Dest' not in values

    def test_date_range_filter_applied(self, app, admin_client):
        from app import db
        from datetime import datetime, timezone
        old_inq = _make_inquiry(db, email='old@example.com', destination='Old Dest')
        old_inq.created_at = datetime(2020, 1, 1, tzinfo=timezone.utc)
        db.session.commit()
        _make_inquiry(db, email='recent@example.com', destination='Recent Dest')
        response = admin_client.get('/admin/inquiries/export?date_from=2024-01-01')
        values = _xlsx_strings(response.data)
        assert 'Recent Dest' in values
        assert 'Old Dest' not in values

    def test_filename_reflects_status_filter(self, app, admin_client):
        response = admin_client.get('/admin/inquiries/export?status=new')
        disposition = response.headers.get('Content-Disposition', '')
        assert 'new' in disposition
        assert '.xlsx' in disposition
